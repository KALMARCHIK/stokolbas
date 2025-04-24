""" Парсер borisovmeat.by """

from os import getcwd
from re import (
    search as re_search,
    findall as re_findall
)
from functools import partial
from asyncio import (
    get_event_loop, 
    gather,
    sleep as async_sleep
)
from concurrent.futures import ThreadPoolExecutor

from openpyxl import load_workbook
from xlrd import open_workbook
from aiohttp import (
    ClientSession,
    ClientTimeout
)
from bs4 import BeautifulSoup as BS

from .common import name_matching


async def parse(parse_data):
    """
    Парсинг сайта, excel файлов и сопоставление полученных данных
    """

    tasks = [parse_site()]
    for f in parse_data['files']:
        tasks.append(parse_file(f))

    results = await gather(*tasks)

    parsed_data = results[0]
    pricelist_data = [x for p in results[1:] for x in p]

    return data_matching(parsed_data, pricelist_data)


def data_matching(parsed_data, pricelist_data):
    """
    Возвращает итоговый набор данных:
    - наименование
    - категория 
    - состав
    - изображение
    - срок годности
    - цена за 1 кг
    - цена за 1 кг от 1 т.
    """

    results = []

    name_matching_threshold = 0.70

    for pd in parsed_data:
        in_quotes = re_findall(r'"(.+?)"', pd['name'])
        if len(in_quotes) > 0:
            sub_name = in_quotes[0].lower()

            for pl in pricelist_data:
                if sub_name in pl['name'].lower():
                    results.append(pl | pd)
                    break
        else:
            # TODO: нужно бы использовать кучу
            cfs = []
            for pl in pricelist_data:
                cf = name_matching(pd['name'], pl['name'])
                if cf >= name_matching_threshold:
                    cfs.append((pl, cf))
            cfs.sort(key=lambda x: x[1])
            if len(cfs) > 0:
                results.append(cfs[-1][0] | pd)

    return results


async def parse_file(file_name):
    """
    Используется для вызова "долгой" блокирующей функции
    в доп процессе
    """

    with ThreadPoolExecutor() as executor:
        loop = get_event_loop()
        return await loop.run_in_executor(
            executor,
            partial(parse_file_blocking, file_name)
        )

def parse_file_blocking(file_name):
    """
    Возвращает список:
    - наименование
    - срок годности
    - цена за 1кг
    - цена за 1кг от 1 т.
    """

    def old_excel_format(sh, ind):
        while ind < sh.nrows:
            v1 = sh.cell_value(rowx=ind, colx=1)
            v2 = sh.cell_value(rowx=ind, colx=2)
            v3 = sh.cell_value(rowx=ind, colx=3)
            v4 = sh.cell_value(rowx=ind, colx=4)
            ind += 1
            yield v1,v2,v3,v4

    def new_excel_format(sh, ind):
        while ind < sh.max_row:
            v1 = sh.cell(row=ind, column=2).value
            v2 = sh.cell(row=ind, column=3).value
            v3 = sh.cell(row=ind, column=4).value
            v4 = sh.cell(row=ind, column=5).value
            ind += 1
            yield v1,v2,v3,v4


    results = []
    __iter = None

    file_name = getcwd() + "/media/" + file_name

    if re_search(r'.xls$', file_name):
        wb = open_workbook(file_name)
        sh = wb.sheet_by_index(0)
        __iter = old_excel_format(sh, 8)
    else:
        wb = load_workbook(file_name, data_only=True)
        sh = wb.worksheets[0]
        __iter = new_excel_format(sh, 9)

    for name, expiration_date, price, price_by_ton in __iter:
        if name is None or expiration_date is None:
            continue

        if isinstance(expiration_date, int):
            expiration_date = float(expiration_date)
        elif not isinstance(expiration_date, float):
            expiration_date = expiration_date.strip().lower()
            if len(name) == 0 and len(expiration_date) == 0:
                break
            if len(expiration_date) == 0:
                continue

            if 'год' in expiration_date:
                try:
                    expiration_date = float(re_findall(r'\d+', expiration_date)[0]) * 365
                except (IndexError, ValueError):
                    expiration_date = 0
            else:
                try:
                    expiration_date = float(expiration_date)
                except ValueError:
                    expiration_date = 0

        if not isinstance(price, float):
            price = 0.0
        if not isinstance(price_by_ton, float):
            price_by_ton = 0.0

        results.append({
            'name': name,
            'expiration_date': expiration_date,
            'price': price,
            'price_by_ton': price_by_ton
        })

    return results


async def parse_site():
    """
    Возвращает список:
    - наименование
    - категория
    - состав
    - ссылку на изображение
    """

    categories_urls = [
        "https://borisovmeat.by/?page_id=533",
        "https://borisovmeat.by/?page_id=562",
        "https://borisovmeat.by/?page_id=629",
        "https://borisovmeat.by/?page_id=584",
        "https://borisovmeat.by/?page_id=598",
        "https://borisovmeat.by/?page_id=577",
        "https://borisovmeat.by/?page_id=569",
        "https://borisovmeat.by/?page_id=548",
    ]


    async def some_sleep(ind, fn):
        await async_sleep(ind * 2)
        return await fn

    async def get_product_data(session, url):
        print('get product data:', url)
        result = {
            'name': '',
            'category': '',
            'composition': '',
            'img': ''
        }


        while True:
            async with session.get(url) as resp:
                if not resp.ok:
                    print('bad response status, repeat')
                    await async_sleep(5)
                    continue
                content = await resp.text()
                break

        soup = BS(content, 'html.parser')
        name_el = soup.find('h2', class_='elementor-heading-title')
        if name_el:
            result['name'] = name_el.text.strip()

        try:
            category = re_findall(r'<p><strong>Категория:<\/strong>(.+?)<', content)[0]
        except IndexError:
            pass
        else:
            result['category'] = category.strip()

        try:
            composition = re_findall(r'<p><strong>Состав продукта:<\/strong>(.+?)<', content)[0]
        except IndexError:
            pass
        else:
            result['composition'] = composition.strip()

        imgs = soup.select('img.attachment-large.size-large')
        if len(imgs) > 0 and imgs[0].has_attr('src'):
            result['img'] = imgs[0]['src']

        return result


    async def processing_categories(session, url):
        async with session.get(url) as resp:
            content = await resp.text()

        soup = BS(content, 'html.parser')
        products_hrefs_els = soup.select(
            'div.elementor-element a.elementor-element.e-con-full.e-transform.e-flex.e-con.e-child'
        )

        tasks = [some_sleep(ind, get_product_data(session, a['href']))
                 for ind, a in enumerate(products_hrefs_els) if a.has_attr('href')]
        return await gather(*tasks)

    async with ClientSession() as session:
        tasks = [some_sleep(ind, processing_categories(session, category_url))
                 for ind, category_url in enumerate(categories_urls)]
        results = [__result for __result_list in await gather(*tasks) for __result in __result_list]

    return results


if __name__ == "__main__":
    data_to_parse = {
        'files': ['./../media/price_lists/ПРАЙС_Борисовский_МК_ЯНВАРЬ_2025.xls'],
        'url': 'https://borisovmeat.by',
    }
    

    from asyncio import run
    from pprint import pprint
    from json import dumps

    #result = run(parse_site(data_to_parse['url']))
    #pprint(result)
    #print(len(result))

    #with open('save.json', 'w', encoding="utf-8") as file:
    #    file.write(dumps(result))

    #exit()

    result = run(parse(data_to_parse))
    pprint(result)
